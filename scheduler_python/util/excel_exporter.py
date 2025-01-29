# scheduler_python/util/excel_exporter.py

import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange
import datetime
import math

class ExcelExporter:
    """
    Python version of the Java ExcelExporter.
    We'll replicate the same style (colors), merges, formulas, etc. 
    using openpyxl.
    """
    def __init__(self, scheduler, people, local_date):
        self.scheduler = scheduler
        self.people = people
        self.local_date = local_date
        self.df_format = "{:.2f}"

    def write_month_to_excel(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Monthly Schedule"

        # styles
        # We replicate the various styles from the Java code:
        def cell_style(fill_color=None, font_color="000000", fill_type="solid"):
            """
            fill_color: a hex or something that openpyxl recognizes, or None
            font_color: hex for the text color
            fill_type: pattern style
            """
            st = openpyxl.styles.NamedStyle()
            if fill_color:
                st.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type=fill_type)
            st.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            st.font = Font(color=font_color)
            return st

        # We'll define some color constants that mirror the Java's IndexedColors approach
        # Java's IndexedColors references:
        #   BLACK, ORANGE, RED, GREY_25_PERCENT, PALE_BLUE, ...
        # We'll approximate them.
        color_black = "000000"
        color_white = "FFFFFF"
        color_grey25 = "D9D9D9"  # approximate
        color_orange = "FFA500"
        color_red = "FF0000"
        color_pale_blue = "D0E3FA"  # approximate
        color_light_yellow = "FFFFCC"  # approximate
        color_aqua = "00FFFF"
        color_light_green = "CCFFCC"
        color_light_turquoise = "CCFFFF"
        color_teal = "008080"
        color_coral = "FF7F50"

        # define cell styles
        top_left_cell_style = cell_style(fill_color="333333", font_color=color_white, fill_type="darkGray")
        header_cell_style = cell_style(fill_color="333333", font_color=color_white)
        header_red_cell_style = cell_style(fill_color="FF0000", font_color=color_white)
        header_orange_cell_style = cell_style(fill_color="FFA500", font_color=color_black)
        wd_style = cell_style(fill_color="000000", font_color="FF0000")  # black fill, red font
        light_grey_style = cell_style(fill_color=color_grey25, font_color=color_black)
        IMS1_style = cell_style(fill_color=color_pale_blue, font_color=color_black)
        IMS2_style = cell_style(fill_color=color_light_yellow, font_color=color_black)
        header_light_green_cell_style = cell_style(fill_color=color_light_green, font_color=color_black)
        header_light_orange_cell_style = cell_style(fill_color=color_aqua, font_color=color_black)
        header_blue_grey_cell_style = cell_style(fill_color=color_grey25, font_color=color_black)
        header_pale_blue_cell_style = cell_style(fill_color=color_pale_blue, font_color=color_black)
        header_pink_cell_style = cell_style(fill_color=color_light_turquoise, font_color=color_black)
        basic_style = cell_style(fill_color=color_teal, font_color=color_black)
        sum_cell_style = cell_style(fill_color=color_coral, font_color=color_black)

        rowNum = 1
        colNum = 1

        # Create header row
        headerRow = sheet.row_dimensions[rowNum]
        # Just set the top-left cell
        cell = sheet.cell(row=rowNum, column=colNum)
        # local_date is a date object; let's do year-month
        cell.value = f"{self.local_date.year} {self.local_date.strftime('%b')}"
        cell.style = top_left_cell_style

        # we track normal weekends
        saturdays = self.scheduler.get_saturdays()
        sundays = self.scheduler.get_sundays()
        normal_weekends = saturdays + sundays
        holidays = self.scheduler.get_holidays()

        # set column width for col 1
        sheet.column_dimensions[get_column_letter(1)].width = 18

        # fill the top row with day numbers
        for i in range(1, self.scheduler.get_num_of_days()+1):
            c = sheet.cell(row=rowNum, column=colNum+i)
            c.value = i
            if i in holidays:
                c.style = header_red_cell_style
            elif i in normal_weekends:
                c.style = header_orange_cell_style
            else:
                c.style = header_cell_style
            sheet.column_dimensions[get_column_letter(colNum+i)].width = 5

        rowNum += 1
        # Now a row with day-of-week abbreviations
        cell = sheet.cell(row=rowNum, column=1)
        cell.style = top_left_cell_style

        for i in range(1, self.scheduler.get_num_of_days()+1):
            c = sheet.cell(row=rowNum, column=colNum+i)
            day_date = datetime.date(self.local_date.year, self.local_date.month, i)
            dow_abr = day_date.strftime('%a').upper()[:3]
            c.value = dow_abr
            if i in holidays:
                c.style = header_red_cell_style
            elif i in normal_weekends:
                c.style = header_orange_cell_style
            else:
                # weekday style
                c.style = wd_style

        rowNum += 1

        # names sorted:
        names = sorted([p.get_name() for p in self.people.get_people().values()])

        # For each name, fill the row with "IMS1"/"IMS2"/"X" for hated, etc.
        for idx, name in enumerate(names):
            places = sorted([d for d, st in self.scheduler.get_scheduled().items() if name in st])
            # gather hated
            p = self.people.get_people()[name]
            hated_days = p.hated_days
            row = rowNum + idx
            # name cell
            c = sheet.cell(row=row, column=1)
            c.value = name
            if (row % 2)==0:
                c.style = header_light_green_cell_style
            else:
                c.style = header_light_orange_cell_style

            # fill day columns
            for i in range(1, self.scheduler.get_num_of_days()+1):
                c2 = sheet.cell(row=row, column=colNum+i)
                if (row % 2)==0:
                    c2.style = header_light_green_cell_style
                else:
                    c2.style = header_light_orange_cell_style

                if i in places:
                    # check if name is FO for that day
                    if i in self.scheduler.get_fo_names():
                        foName = self.scheduler.get_fo_names()[i]
                        if foName == name:
                            c2.value = "IMS1"
                            c2.style = IMS1_style
                        else:
                            c2.value = "IMS2"
                            c2.style = IMS2_style
                    else:
                        c2.value = "NULL"
                        # or some default style
                elif i in hated_days:
                    c2.value = "X"
                    c2.style = light_grey_style
                else:
                    # if it's a holiday => red
                    if i in holidays:
                        c2.style = header_red_cell_style
                    elif i in normal_weekends:
                        c2.style = header_orange_cell_style
                    else:
                        # just keep the style from above
                        pass

        # skip a few rows
        rowNum = rowNum + len(names) + 3

        # Now the "SUMMARIZE" table
        row = sheet.row_dimensions[rowNum]
        for i in range(1, 10):
            c = sheet.cell(row=rowNum, column=i)
            # just placeholders
        c2 = sheet.cell(row=rowNum, column=10)
        c2.value = f"{self.local_date.year} SUMMARIZE"
        c2.style = header_light_orange_cell_style
        # we'll merge columns 10..15
        sheet.merge_cells(start_row=rowNum, start_column=10, end_row=rowNum, end_column=15)

        # next cells: "up to ...
        colX = 10 + 6
        c3 = sheet.cell(row=rowNum, column=colX)
        end_of_month = self.local_date.replace(day=self.scheduler.get_num_of_days())
        c3.value = f"up to {end_of_month.strftime('%b')} {end_of_month.day}"
        c3.style = header_light_orange_cell_style
        sheet.merge_cells(start_row=rowNum, start_column=colX, end_row=rowNum, end_column=colX+4)

        rowNum+=1
        # create a row of headers for the summarize table
        r = rowNum
        col = 1

        def setval_and_style(crow, ccol, val, sty):
            cc = sheet.cell(row=crow, column=ccol)
            cc.value = val
            cc.style = sty

        setval_and_style(r,col,"",header_light_green_cell_style)
        col+=1
        setval_and_style(r,col,"IMS1\nWE",header_blue_grey_cell_style); col+=1
        setval_and_style(r,col,"IMS1\nWD",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"IMS2\nWE",header_blue_grey_cell_style); col+=1
        setval_and_style(r,col,"IMS2\nWD",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"NH",header_pink_cell_style); col+=1
        setval_and_style(r,col,"WE",header_pink_cell_style); col+=1
        setval_and_style(r,col,"WD",header_pink_cell_style); col+=1
        setval_and_style(r,col,"ALL",header_pink_cell_style); col+=1
        setval_and_style(r,col,"MON-\nTHU",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"FR",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"SA",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"SU",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"WE",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"NH",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"ALL",header_pale_blue_cell_style); col+=1
        setval_and_style(r,col,"MON-\nTHU",header_blue_grey_cell_style); col+=1
        setval_and_style(r,col,"FRI",header_blue_grey_cell_style); col+=1
        setval_and_style(r,col,"SAT",header_blue_grey_cell_style); col+=1
        setval_and_style(r,col,"SUN",header_blue_grey_cell_style); col+=1
        setval_and_style(r,col,"NH",header_blue_grey_cell_style); col+=1
        setval_and_style(r,col,"STANDBY\n WE                 WD               SUM",header_pink_cell_style); col+=1

        rowNum+=1

        # We'll replicate the big loop from Java that calculates all the partial sums
        # For the day-of-week classification, we'll gather them in Python:
        #   fridays, saturdays, sundays => from the scheduler
        def day_of_week(d):
            # 0=Mon,...6=Sun
            dd = datetime.date(self.local_date.year, self.local_date.month, d)
            return dd.weekday()

        # gather fridays
        fridays = [i for i in range(1,self.scheduler.get_num_of_days()+1) if day_of_week(i)==4]
        saturdays = [i for i in range(1,self.scheduler.get_num_of_days()+1) if day_of_week(i)==5]
        sundays = [i for i in range(1,self.scheduler.get_num_of_days()+1) if day_of_week(i)==6]

        # For each name:
        rindex = 0
        for name in names:
            rowIdx = rowNum + rindex
            rindex+=1
            c = sheet.cell(row=rowIdx, column=1)
            c.value = name
            if (rowIdx%2)==0:
                c.style = header_light_green_cell_style
            else:
                c.style = header_light_orange_cell_style

            scheduled_days = [d for d, st in self.scheduler.get_scheduled().items() if name in st]
            # separate out holiday vs not holiday
            # holiday + friday? holiday + saturday? etc
            count_fridays = sum(1 for d in scheduled_days if d in fridays and d not in holidays)
            count_saturdays = sum(1 for d in scheduled_days if d in saturdays and d not in holidays)
            count_sundays = sum(1 for d in scheduled_days if d in sundays and d not in holidays)
            count_fri_holidays = sum(1 for d in scheduled_days if (d in fridays) and (d in holidays))
            count_sat_holidays = sum(1 for d in scheduled_days if (d in saturdays) and (d in holidays))
            count_sun_holidays = sum(1 for d in scheduled_days if (d in sundays) and (d in holidays))
            count_weekday_holidays = sum(1 for d in scheduled_days
                                         if (d in holidays) and (d not in fridays) and (d not in saturdays) and (d not in sundays))
            count_holidays = count_fri_holidays + count_sat_holidays + count_sun_holidays + count_weekday_holidays

            # how many times is name FO vs BO
            fo_scheduled = [dd for dd, f in self.scheduler.get_fo_names().items() if f == name]
            ims1_weekend = sum(1 for dd in fo_scheduled if dd in saturdays or dd in sundays)
            ims1_weekday = len(fo_scheduled) - ims1_weekend
            # who else is scheduled on those days => that is IMS2
            bo_scheduled = [d for d in scheduled_days if d not in fo_scheduled]
            ims2_weekend = sum(1 for dd in bo_scheduled if dd in saturdays or dd in sundays)
            ims2_weekday = len(bo_scheduled) - ims2_weekend

            colIdx = 2
            def setnum(x):
                cc = sheet.cell(row=rowIdx, column=colIdx)
                cc.value = x
                if (rowIdx%2)==0:
                    cc.style = header_light_green_cell_style
                else:
                    cc.style = header_light_orange_cell_style
            setnum(ims1_weekend); colIdx+=1
            setnum(ims1_weekday); colIdx+=1
            setnum(ims2_weekend); colIdx+=1
            setnum(ims2_weekday); colIdx+=1
            setnum(count_holidays); colIdx+=1
            setnum(ims1_weekend+ims2_weekend); colIdx+=1
            setnum(ims1_weekday+ims2_weekday); colIdx+=1
            setnum(len(scheduled_days)); colIdx+=1

            # For the second half, the Java code does DB queries to sum previous months. We'll replicate that if you want, 
            # or we can do the short version. We'll replicate with KeyValueStore:
            kv = self.scheduler.get_key_value_store()
            y = self.local_date.year
            m = self.local_date.month
            allAll = kv.sum(name, y, m, "ALL") + len(scheduled_days)
            frAll = kv.sum(name, y, m, "FR") + count_fridays
            suAll = kv.sum(name, y, m, "SU") + count_sundays
            weAll = kv.sum(name, y, m, "WE") + (count_saturdays + count_sundays)
            nhwdAll = kv.sum(name, y, m, "NHWD") + count_weekday_holidays
            nhfrAll = kv.sum(name, y, m, "NHFR") + count_fri_holidays
            nhsaAll = kv.sum(name, y, m, "NHSA") + count_sat_holidays
            nhsuAll = kv.sum(name, y, m, "NHSU") + count_sun_holidays
            nhAll = nhwdAll + nhfrAll + nhsaAll + nhsuAll
            mon2thuAll = allAll - (weAll + nhsaAll + nhsuAll) - (frAll + nhfrAll) - nhwdAll
            saAll = weAll - suAll

            def setnumX(x):
                cc = sheet.cell(row=rowIdx, column=colIdx)
                cc.value = x
                if (rowIdx%2)==0:
                    cc.style = header_light_green_cell_style
                else:
                    cc.style = header_light_orange_cell_style

            setnumX(mon2thuAll); colIdx+=1
            setnumX(frAll); colIdx+=1
            setnumX(saAll); colIdx+=1
            setnumX(suAll); colIdx+=1
            setnumX( (weAll + nhsaAll + nhsuAll) ); colIdx+=1
            setnumX(nhAll); colIdx+=1
            setnumX(allAll); colIdx+=1

            # Now the local counts for mon-thu, fri, sat, sun, holidays in *this* month:
            mon_to_thu = [d for d in scheduled_days if (d not in saturdays) and (d not in sundays) and (d not in fridays) and (d not in holidays)]
            setnumX(len(mon_to_thu)); colIdx+=1
            only_fr = [d for d in scheduled_days if (d in fridays) and (d not in holidays)]
            setnumX(len(only_fr)); colIdx+=1
            setnumX(count_saturdays); colIdx+=1
            setnumX(count_sundays); colIdx+=1
            setnumX(count_holidays); colIdx+=1

            # The "weekendSBNum" and "weekdaySBNum" the Java code does some big formula:
            # weekendSBNum = 3.6*frAll + 9.6*saAll + 6*suAll + 9.6*(nhfrAll+nhsaAll+nhsuAll+nhwdAll?)
            # Actually see the code carefully. We'll replicate:
            weekendSBNum = 3.6*frAll + 9.6*saAll + 6*suAll + 9.6*(nhfrAll + nhsaAll + nhsuAll + nhwdAll)
            weekdaySBNum = 3.2*mon2thuAll + 1.4*frAll + 1.8*suAll

            def setnumF(x):
                cc = sheet.cell(row=rowIdx, column=colIdx)
                cc.value = float(self.df_format.format(x))
                if (rowIdx%2)==0:
                    cc.style = header_light_green_cell_style
                else:
                    cc.style = header_light_orange_cell_style

            setnumF(weekendSBNum)
            colIdx+=2 # skip 1 col, merge
            setnumF(weekdaySBNum)
            colIdx+=2 # skip 1 col, merge
            totalSB = weekendSBNum + weekdaySBNum
            setnumF(totalSB)
            colIdx+=2 # skip 1 col, merge

            # Finally, store in DB the updated values
            kv.write_data(name, y, m, "ALL", len(scheduled_days))
            kv.write_data(name, y, m, "FR", count_fridays)
            kv.write_data(name, y, m, "SU", count_sundays)
            kv.write_data(name, y, m, "WE", (count_saturdays+count_sundays))
            kv.write_data(name, y, m, "NHWD", count_weekday_holidays)
            kv.write_data(name, y, m, "NHFR", count_fri_holidays)
            kv.write_data(name, y, m, "NHSA", count_sat_holidays)
            kv.write_data(name, y, m, "NHSU", count_sun_holidays)

        # Summation row
        total_row = rowNum + len(names)
        c = sheet.cell(row=total_row, column=1)
        for i in range(23):
            # We'll place a formula: =SUM(...)
            # We have to figure out the letter for the column
            colLetter = get_column_letter(i+2)
            # the data range is from rowNum..(rowNum+len(names)-1)
            firstDataRow = rowNum
            lastDataRow = rowNum + len(names)-1
            formula = f"=SUM({colLetter}{firstDataRow}:{colLetter}{lastDataRow})"
            sumCell = sheet.cell(row=total_row, column=i+2)
            sumCell.value = formula
            sumCell.style = sum_cell_style
            # The Java code merges if i>19? We'll skip the merges for simplicity. 
            # If you want to replicate exactly, you'd have to replicate those merges.

        # printAll
        self.scheduler.get_key_value_store().print_all(self.local_date.year)
        # close DB
        try:
            self.scheduler.get_key_value_store().close()
        except:
            pass

        # Save
        fileNameExcel = f"schedule-{self.local_date.year}-{self.local_date.month}.xlsx"
        wb.save(fileNameExcel)
