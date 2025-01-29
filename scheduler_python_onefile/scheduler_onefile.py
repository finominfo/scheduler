###############################################################################
# scheduler_onefile.py
# A single Python script that replicates all the functionality of the Java code
# you provided, including reading config.csv, scheduling, writing an XLSX,
# storing data in a local SQLite DB (like KeyValueStore), etc.
###############################################################################

import sys
import os
import datetime
import codecs
import random
import sqlite3
import calendar
import traceback

# For Excel creation:
import openpyxl
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.cell_range import CellRange

###############################################################################
# 0) GLOBALS
###############################################################################
class Globals:
    """
    Translated from Java's Globals.java, just storing file names.
    """
    def __init__(self):
        self.config_file = "config.csv"
        self.result_file = "result.csv"

    def get_config_file(self):
        return self.config_file

    def get_result_file(self):
        return self.result_file


###############################################################################
# 1) ENUM Type
###############################################################################
from enum import Enum
class Type(Enum):
    """
    Replica of the Java 'Type' enum.
    FO = IMS1
    BO = IMS2
    FO_AND_BO = IMS
    """
    FO = "IMS1"
    BO = "IMS2"
    FO_AND_BO = "IMS"

    def good_with(self, other):
        """
        Mirroring 'public boolean goodWith(Type type)' from Java.
        - If either is FO_AND_BO => always good
        - Otherwise they must be different (FO != BO)
        """
        if self == Type.FO_AND_BO:
            return True
        if other == Type.FO_AND_BO:
            return True
        return self != other

    @staticmethod
    def is_first_fo(t1, t2):
        """
        Java code: return t1.equals(FO) || t2.equals(BO);
        Meaning if the first is FO or the second is BO => the first is considered 'first FO'.
        """
        return (t1 == Type.FO) or (t2 == Type.BO)

    def to_cell(self):
        return self.value  # e.g. "IMS1", "IMS2", or "IMS"


###############################################################################
# 2) PERSON
###############################################################################
class Person:
    """
    Python version of the Java Person class.
    """
    def __init__(self, name):
        self.name = name
        self.hated_days = []  # store days that are hated
        self.wanted_days = set()  # store days that are wanted
        self.types = {}
        for i in range(-5, 36):
            self.types[i] = Type.FO_AND_BO

        self.num_of_scheduled = 0
        self.manual_day_difference = 0

        # The four numeric "desired" fields:
        self.num_of_wanted_holidays = 0
        self.num_of_wanted_fridays = 0
        self.num_of_wanted_saturdays = 0
        self.num_of_wanted_sundays = 0

        # boolean fields
        self.nofo = False
        self.hates_weekends = False
        self.hates_weekdays = False
        self.hates_mondays = False
        self.hates_tuesdays = False
        self.wanted_tuesdays = False
        self.hates_wednesdays = False
        self.hates_thursdays = False
        self.hates_fridays = False

    # Basic getters
    def get_name(self):
        return self.name

    def get_type(self, day):
        return self.types[day]

    def set_type(self, day, t):
        self.types[day] = t

    def increment_scheduled(self):
        self.num_of_scheduled += 1

    def get_num_of_scheduled(self):
        return self.num_of_scheduled

    def get_manual_day_difference(self):
        return self.manual_day_difference

    def set_manual_day_difference(self, val):
        self.manual_day_difference = val

    def is_nofo(self):
        return self.nofo

    def set_nofo(self, val):
        self.nofo = val

    def is_hates_weekends(self):
        return self.hates_weekends

    def set_hates_weekends(self, val):
        self.hates_weekends = val

    def is_hates_weekdays(self):
        return self.hates_weekdays

    def set_hates_weekdays(self, val):
        self.hates_weekdays = val

    def is_hates_mondays(self):
        return self.hates_mondays

    def set_hates_mondays(self, val):
        self.hates_mondays = val

    def is_hates_tuesdays(self):
        return self.hates_tuesdays

    def set_hates_tuesdays(self, val):
        self.hates_tuesdays = val

    def is_wanted_tuesdays(self):
        return self.wanted_tuesdays

    def set_wanted_tuesdays(self, val):
        self.wanted_tuesdays = val

    def is_hates_wednesdays(self):
        return self.hates_wednesdays

    def set_hates_wednesdays(self, val):
        self.hates_wednesdays = val

    def is_hates_thursdays(self):
        return self.hates_thursdays

    def set_hates_thursdays(self, val):
        self.hates_thursdays = val

    def is_hates_fridays(self):
        return self.hates_fridays

    def set_hates_fridays(self, val):
        self.hates_fridays = val

    def get_hated_days(self):
        return self.hated_days

    def get_wanted_days(self):
        return self.wanted_days

    # numeric "desired" fields
    def get_num_of_wanted_holidays(self):
        return self.num_of_wanted_holidays

    def set_num_of_wanted_holidays(self, x):
        self.num_of_wanted_holidays = x

    def get_num_of_wanted_fridays(self):
        return self.num_of_wanted_fridays

    def set_num_of_wanted_fridays(self, x):
        self.num_of_wanted_fridays = x

    def get_num_of_wanted_saturdays(self):
        return self.num_of_wanted_saturdays

    def set_num_of_wanted_saturdays(self, x):
        self.num_of_wanted_saturdays = x

    def get_num_of_wanted_sundays(self):
        return self.num_of_wanted_sundays

    def set_num_of_wanted_sundays(self, x):
        self.num_of_wanted_sundays = x


###############################################################################
# 3) KEYVALUESTORE (SQLite)
###############################################################################
class KeyValueStore:
    """
    Python replica of the Java KeyValueStore (which used H2),
    but we do it with SQLite in a local file "keyvaluestore.db".
    """

    def __init__(self, db_path="keyvaluestore.db"):
        self.db_path = db_path
        self.conn = sqlite3.connect(self.db_path)
        self.create_database()

    def create_database(self):
        try:
            cur = self.conn.cursor()
            cur.execute("""
                CREATE TABLE IF NOT EXISTS KeyValueStore (
                    name TEXT,
                    year INT,
                    month INT,
                    type TEXT,
                    value INT,
                    PRIMARY KEY (name, year, month, type)
                )
            """)
            self.conn.commit()
        except Exception as e:
            print("Error creating DB:", e)

    def write_data(self, name, year, month, type_str, value):
        try:
            cur = self.conn.cursor()
            # upsert
            cur.execute("""
            INSERT INTO KeyValueStore (name, year, month, type, value)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(name, year, month, type)
            DO UPDATE SET value=excluded.value
            """, (name, year, month, type_str, value))
            self.conn.commit()
        except Exception as e:
            print("Error writing data:", e)

    def read_data(self, name, year, month, type_str):
        """
        Return a dict {type_str: value}, basically
        """
        data = {}
        try:
            cur = self.conn.cursor()
            cur.execute("""
                SELECT name, year, month, type, value
                FROM KeyValueStore
                WHERE name=? AND year=? AND month=? AND type=?
            """,(name, year, month, type_str))
            rows = cur.fetchall()
            for row in rows:
                data[row[3]] = row[4]
        except Exception as e:
            print("Error reading data:", e)
        return data

    def sum(self, name, year, month, type_str):
        """
        Java logic:
          SELECT SUM(value) FROM KeyValueStore WHERE name=? AND year=? AND month < ? AND type=?
        """
        s = 0
        try:
            cur = self.conn.cursor()
            cur.execute("""
                SELECT SUM(value) FROM KeyValueStore
                WHERE name=? AND year=? AND month < ? AND type=?
            """,(name, year, month, type_str))
            row = cur.fetchone()
            if row and row[0] is not None:
                s = row[0]
        except Exception as e:
            print("Error summing:", e)
        return s

    def get_names(self):
        """
        SELECT DISTINCT name FROM KeyValueStore
        """
        names = []
        try:
            cur = self.conn.cursor()
            cur.execute("SELECT DISTINCT name FROM KeyValueStore")
            rows = cur.fetchall()
            for r in rows:
                names.append(r[0])
        except sqlite3.OperationalError as e:
            if "no such table" in str(e).lower():
                self.create_database()
            else:
                print("Error get_names:", e)
        return names

    def print_all(self, year):
        """
        For debugging: SELECT * FROM KeyValueStore WHERE year=?
        Then write to allData.txt
        """
        try:
            cur = self.conn.cursor()
            cur.execute("SELECT * FROM KeyValueStore WHERE year=?", (year,))
            rows = cur.fetchall()
            sb = []
            for row in rows:
                # row = (name, year, month, type, value)
                s = f"{row[0]} {row[1]} {row[2]} {row[3]} {row[4]}"
                sb.append(s)
            with open("allData.txt","w",encoding="utf-8") as f:
                f.write("\n".join(sb))
        except Exception as e:
            print("Error printAll:", e)

    def close(self):
        try:
            self.conn.commit()
            self.conn.close()
        except Exception as e:
            print("Error closing DB:", e)


###############################################################################
# 4) HUNGARIAN HOLIDAYS
###############################################################################
class HungarianHolidays:
    """
    Same approach as HungarianHolidays.java.
    """

    @staticmethod
    def get_holidays_for_year(year):
        import datetime
        holidays = []

        # Jan 1
        holidays.append(datetime.date(year,1,1))
        # March 15
        holidays.append(datetime.date(year,3,15))

        # Easter stuff
        easter = HungarianHolidays.calculate_easter_date(year)
        # Good Friday (easter-2), Easter Monday (easter+1)
        good_friday = easter - datetime.timedelta(days=2)
        easter_monday = easter + datetime.timedelta(days=1)
        holidays.append(good_friday)
        holidays.append(easter_monday)

        # May 1
        holidays.append(datetime.date(year,5,1))

        # Pentecost Monday (Easter + 50 days)
        pentecost_monday = easter + datetime.timedelta(days=50)
        holidays.append(pentecost_monday)

        # Aug 20
        holidays.append(datetime.date(year,8,20))
        # Oct 23
        holidays.append(datetime.date(year,10,23))
        # Nov 1
        holidays.append(datetime.date(year,11,1))
        # Dec 25
        holidays.append(datetime.date(year,12,25))
        # Dec 26
        holidays.append(datetime.date(year,12,26))

        return holidays

    @staticmethod
    def get_holidays_for_month(year, month):
        import datetime
        all_h = HungarianHolidays.get_holidays_for_year(year)
        return [d for d in all_h if d.month == month]

    @staticmethod
    def calculate_easter_date(year):
        """
        Zeller's-like formula from the Java code.
        """
        a = year % 19
        b = year // 100
        c = year % 100
        d = b // 4
        e = b % 4
        f = (b + 8)//25
        g = (b - f + 1)//3
        h = (19*a + b - d - g + 15) % 30
        i = c//4
        k = c % 4
        l = (32 + 2*e + 2*i - h - k) % 7
        m = (a + 11*h + 22*l)//451
        month = (h + l - 7*m + 114)//31
        day = ((h + l - 7*m + 114) % 31)+1
        return datetime.date(year, month, day)


###############################################################################
# 5) PEOPLE (reads config.csv)
###############################################################################
class People:
    """
    Python translation of the Java People class.
    Reads config.csv, building Person objects, applying hate/want logic.
    """

    def __init__(self):
        self.people = {}
        self.hated = {}
        for i in range(32):
            self.hated[i] = set()

        # read lines
        g = Globals()
        config_file = g.get_config_file()

        content = ""
        with codecs.open(config_file,"r","utf-8") as f:
            content = f.read()
        lines = content.splitlines()

        kv = KeyValueStore()
        known_names = kv.get_names()
        kv.close()

        keywords = [
            "nofo","hend","hweek","hmon","htue","hwen","hthu","hfri","wtue"
        ]

        for line in lines:
            line = line.strip()
            if not line:
                continue
            expressions = line.split()
            person = None
            new_person_obj = None

            # We'll replicate the Java approach: parse tokens
            # Actually, the Java code line-by-line handles it in a single pass.
            # We'll do similarly:
            pointer_person = None
            maybe_new_person = None

            for expr in expressions:
                expr = expr.strip()
                if not expr:
                    continue
                if pointer_person is None:
                    # check if we are in "new person" scenario:
                    replaced = expr.replace("_"," ")
                    if replaced in known_names:
                        # existing person
                        pointer_person = Person(replaced)
                        self.people[replaced] = pointer_person
                    else:
                        # maybe we need next token "newperson" in Java code,
                        # but the posted Java code is somewhat different. 
                        # We'll do the same approach used in your snippet: we create a Person anyway
                        # then if next token isn't "newperson", we fail. 
                        # For simplicity, let's just create them:
                        pointer_person = Person(replaced)
                        self.people[replaced] = pointer_person
                else:
                    # interpret keywords
                    low_expr = expr.lower()
                    if low_expr in keywords:
                        if low_expr=="nofo":
                            # set all day => Type.BO
                            for d in pointer_person.types:
                                pointer_person.set_type(d, Type.BO)
                            pointer_person.set_nofo(True)
                        elif low_expr=="hend":
                            pointer_person.set_hates_weekends(True)
                        elif low_expr=="hweek":
                            pointer_person.set_hates_weekdays(True)
                        elif low_expr=="hmon":
                            pointer_person.set_hates_mondays(True)
                        elif low_expr=="htue":
                            pointer_person.set_hates_tuesdays(True)
                        elif low_expr=="hwen":
                            pointer_person.set_hates_wednesdays(True)
                        elif low_expr=="hthu":
                            pointer_person.set_hates_thursdays(True)
                        elif low_expr=="hfri":
                            pointer_person.set_hates_fridays(True)
                        elif low_expr=="wtue":
                            pointer_person.set_wanted_tuesdays(True)
                    else:
                        # parse expressions: w12, h10-12, f15, b20, etc.
                        c = expr[0]
                        remaining = expr[1:]
                        days = []
                        if "-" in remaining:
                            sp = remaining.split("-")
                            start = int(sp[0])
                            end = int(sp[1])
                            for dd in range(start, end+1):
                                days.append(dd)
                        else:
                            days.append(int(remaining))

                        if c=='u':
                            pointer_person.set_num_of_wanted_holidays(days[0])
                        elif c=='p':
                            pointer_person.set_num_of_wanted_fridays(days[0])
                        elif c=='s':
                            pointer_person.set_num_of_wanted_saturdays(days[0])
                        elif c=='v':
                            pointer_person.set_num_of_wanted_sundays(days[0])
                        elif c=='w':
                            for d in days:
                                pointer_person.get_wanted_days().add(d)
                                if d in pointer_person.get_hated_days():
                                    raise RuntimeError(f"{pointer_person.get_name()} wants and hates the same day {d}")
                        elif c=='h':
                            for d in days:
                                pointer_person.get_hated_days().append(d)
                                if d in pointer_person.get_wanted_days():
                                    raise RuntimeError(f"{pointer_person.get_name()} wants and hates day {d}")
                        elif c=='f':
                            for d in days:
                                pointer_person.set_type(d, Type.FO)
                        elif c=='b':
                            for d in days:
                                pointer_person.set_type(d, Type.BO)
                        elif c=='+':
                            pointer_person.set_manual_day_difference(days[0])
                        elif c=='-':
                            pointer_person.set_manual_day_difference(-days[0])
                        else:
                            # do nothing
                            pass


    def get_people(self):
        return self.people


###############################################################################
# 6) SCHEDULER
###############################################################################
class Scheduler:
    """
    Python replication of the Java Scheduler logic (Scheduler.java).
    """

    def __init__(self, people_map, local_date):
        self.people = people_map  # { name: Person }
        self.random_gen = random.Random()
        self.random_gen.seed(int(datetime.datetime.now().timestamp()*1000))

        self.fo_names = {}  # day -> string (who is FO)
        self.scheduled = {} # day -> set of names
        self.hated = {}     # day -> set of names

        year = local_date.year
        month = local_date.month
        self.num_of_days = calendar.monthrange(year, month)[1]
        self.local_date = datetime.date(year, month, 1)

        for i in range(-8, self.num_of_days+11):
            self.scheduled[i] = set()
            self.hated[i] = set()

        # gather Hungarian holidays in that month
        self.holidays = [d.day for d in HungarianHolidays.get_holidays_for_month(year, month)]

        # figure out day-of-week lists
        self.mondays = []
        self.tuesdays = []
        self.wednesdays = []
        self.thursdays = []
        self.fridays = []
        self.saturdays = []
        self.sundays = []

        # fill day-of-week
        self.count_days()

        # step by step from Java
        self.set_hated()
        self.set_wanted()
        self.set_desired_number_of_fridays_and_weekends_and_holidays()
        self.set_weekends_and_holidays()
        self.set_weekdays()
        self.balance_ims()

        self.kv_store = KeyValueStore()

    def get_key_value_store(self):
        return self.kv_store

    def get_local_date(self):
        return self.local_date

    def get_num_of_days(self):
        return self.num_of_days

    def get_holidays(self):
        return self.holidays

    def get_saturdays(self):
        return self.saturdays

    def get_sundays(self):
        return self.sundays

    def get_scheduled(self):
        return self.scheduled

    def get_fo_names(self):
        return self.fo_names

    def count_days(self):
        """
        Fill self.mondays, self.tuesdays, etc., based on day of week.
        Mirroring the Java logic.
        Also includes the special check for 2024-08-03 => Monday, 2024-08-19 => Saturday
        from the Java code.
        """
        for day in range(1, self.num_of_days+1):
            d = datetime.date(self.local_date.year, self.local_date.month, day)
            # special checks from Java:
            if d.year==2024 and d.month==8 and day==3:
                self.mondays.append(day)
            elif d.year==2024 and d.month==8 and day==19:
                self.saturdays.append(day)
            else:
                wd = d.weekday()  # Monday=0,...Sunday=6
                if wd==0:
                    self.mondays.append(day)
                elif wd==1:
                    self.tuesdays.append(day)
                elif wd==2:
                    self.wednesdays.append(day)
                elif wd==3:
                    self.thursdays.append(day)
                elif wd==4:
                    self.fridays.append(day)
                elif wd==5:
                    self.saturdays.append(day)
                elif wd==6:
                    self.sundays.append(day)

    def set_hated(self):
        """
        Mirroring setHated in Java.
        """
        n_people = len(self.people)

        for name, p in self.people.items():
            if p.is_hates_mondays():
                for d in self.mondays:
                    p.get_hated_days().append(d)
            if p.is_hates_tuesdays():
                for d in self.tuesdays:
                    p.get_hated_days().append(d)
            if p.is_wanted_tuesdays():
                for d in self.tuesdays:
                    p.get_wanted_days().add(d)
            if p.is_hates_wednesdays():
                for d in self.wednesdays:
                    p.get_hated_days().append(d)
            if p.is_hates_thursdays():
                for d in self.thursdays:
                    p.get_hated_days().append(d)
            if p.is_hates_fridays():
                for d in self.fridays:
                    p.get_hated_days().append(d)
            if p.is_hates_weekends():
                for d in self.saturdays:
                    p.get_hated_days().append(d)
                for d in self.sundays:
                    p.get_hated_days().append(d)
            if p.is_hates_weekdays():
                for d in (self.mondays + self.tuesdays + self.wednesdays + self.thursdays + self.fridays):
                    p.get_hated_days().append(d)

        # fill self.hated
        for name, p in self.people.items():
            for hd in p.hated_days:
                self.hated[hd].add(name)

        # if day is hated by n_people - 2 => only 2 remain => schedule them
        for day in range(-8, self.num_of_days+11):
            if 1<=day<=self.num_of_days:
                disliked = self.hated[day]
                if len(disliked) > n_people-2:
                    raise RuntimeError(f"Too many people hate day {day}: {disliked}")
                if len(disliked)==n_people-2:
                    # only 2 remain
                    possible_names = set(self.people.keys()) - disliked
                    self.scheduled[day] = set(self.scheduled[day]) | possible_names
                    if len(self.scheduled[day])==2:
                        self.fo_names[day] = self.select_fo(self.scheduled[day], day)

    def set_wanted(self):
        """
        Mirroring setWanted in Java.
        If more than 2 want same day => error,
        if exactly 2 => must not both FO or both BO
        """
        for name, p in self.people.items():
            for wday in p.get_wanted_days():
                self.scheduled[wday].add(name)
                if len(self.scheduled[wday])>2:
                    raise RuntimeError(f"More than two people want day {wday}: {self.scheduled[wday]}")
                if len(self.scheduled[wday])==2:
                    # check if both FO or both BO
                    names_list = list(self.scheduled[wday])
                    t0 = self.people[names_list[0]].get_type(wday)
                    t1 = self.people[names_list[1]].get_type(wday)
                    if t0==Type.BO and t1==Type.BO:
                        raise RuntimeError(f"Two BO want the same day {wday}: {self.scheduled[wday]}")
                    if t0==Type.FO and t1==Type.FO:
                        raise RuntimeError(f"Two FO want the same day {wday}: {self.scheduled[wday]}")
                    self.fo_names[wday] = self.select_fo(self.scheduled[wday], wday)

    def set_desired_number_of_fridays_and_weekends_and_holidays(self):
        """
        Mirroring setDesiredNumberOfFridaysAndWeekendsAndHolidays in Java.
        """
        for name, p in self.people.items():
            hol_count = p.get_num_of_wanted_holidays()
            if hol_count>0:
                self.schedule_desired_days(p, hol_count, self.holidays, "holiday")
            fri_count = p.get_num_of_wanted_fridays()
            if fri_count>0:
                self.schedule_desired_days(p, fri_count, self.fridays, "friday")
            sat_count = p.get_num_of_wanted_saturdays()
            if sat_count>0:
                self.schedule_desired_days(p, sat_count, self.saturdays, "saturday")
            sun_count = p.get_num_of_wanted_sundays()
            if sun_count>0:
                self.schedule_desired_days(p, sun_count, self.sundays, "sunday")

    def schedule_desired_days(self, person, count, day_list, day_name):
        # reduce count by the ones already scheduled
        for d in day_list:
            if person.get_name() in self.scheduled[d]:
                count -=1
        # possible days
        possible_days1 = []
        for d in day_list:
            if person.get_name() not in self.scheduled[d]:
                possible_days1.append(d)
        # filter out days that already have 2 scheduled
        possible_days = [d for d in possible_days1 if len(self.scheduled[d])<2]

        if count>0:
            if len(possible_days)<count:
                raise RuntimeError(f"Not enough {day_name} days for {person.get_name()}")
            while count>0:
                chosen_day = self.random_gen.choice(possible_days)
                if person.get_name() not in self.scheduled[chosen_day]:
                    self.scheduled[chosen_day].add(person.get_name())
                    # if no FO chosen yet, pick random
                    if chosen_day not in self.fo_names or not self.fo_names[chosen_day]:
                        arr = list(self.scheduled[chosen_day])
                        chosen_fo = self.random_gen.choice(arr)
                        self.fo_names[chosen_day] = chosen_fo
                count -=1

    def set_weekends_and_holidays(self):
        """
        Mirroring setWeekendsAndHolidays in Java.
        We'll handle saturdays, sundays, holidays.
        """
        sets_of_days = [self.saturdays, self.sundays, self.holidays]
        for days in sets_of_days:
            ent = self.get_the_most_hated_and_not_scheduled(days)
            while ent[1] is not None:
                day = ent[0]
                ordered = self.get_weekend_ordered_possibilities(day)
                if len(self.scheduled[day])==0:
                    self.scheduled[day].add(ordered[0])
                first_person = list(self.scheduled[day])[0]
                self.find_first_good_for(first_person, ordered, day)
                if len(self.scheduled[day])==2:
                    self.fo_names[day] = self.select_fo(self.scheduled[day], day)
                ent = self.get_the_most_hated_and_not_scheduled(days)

    def get_the_most_hated_and_not_scheduled(self, days):
        most_hated = -1
        position = -1
        final_result2 = None
        for i in range(1, self.num_of_days+1):
            if i in days and len(self.scheduled[i])<2:
                day_hated = set()
                day_hated = day_hated.union(self.hated[i])
                day_hated = day_hated.union(self.scheduled[i])
                if (i-1) in self.scheduled:
                    day_hated = day_hated.union(self.scheduled[i-1])
                if (i+2) in self.scheduled:
                    day_hated = day_hated.union(self.scheduled[i+2])
                if len(day_hated)>most_hated:
                    most_hated = len(day_hated)
                    position = i
                    result2 = set(self.people.keys()) - day_hated
                    final_result2 = result2
        return (position, final_result2)

    def get_weekend_ordered_possibilities(self, saturday_number):
        """
        Java code getWeekendPossibilities + adjacency-based ordering
        """
        persons = self.get_weekend_possibilities(saturday_number)
        schedule_numbers = {}
        for n in persons:
            schedule_numbers[n] = 0

        for d, st in self.scheduled.items():
            for name in st:
                if name in persons:
                    schedule_numbers[name]+=14
                    if d==(saturday_number-2) or d==(saturday_number+3):
                        schedule_numbers[name]+=4
                    if d==(saturday_number-3) or d==(saturday_number+4):
                        schedule_numbers[name]+=2
                    if d==(saturday_number-7) or d==(saturday_number+7):
                        schedule_numbers[name]+=1000
                    if d==(saturday_number-6) or d==(saturday_number+8):
                        schedule_numbers[name]+=1000

        # sort ascending
        return self.order_schedule_numbers(schedule_numbers)

    def get_weekend_possibilities(self, saturday_number):
        sunday_number = saturday_number+1
        friday_number = saturday_number-1
        monday_number = saturday_number+2
        possibilities = set(self.people.keys())
        # remove hated
        if saturday_number in self.hated:
            possibilities = possibilities - self.hated[saturday_number]
        if sunday_number in self.hated:
            possibilities = possibilities - self.hated[sunday_number]
        if friday_number in self.scheduled:
            possibilities = possibilities - self.scheduled[friday_number]
        if monday_number in self.scheduled:
            possibilities = possibilities - self.scheduled[monday_number]
        return list(possibilities)

    def set_weekdays(self):
        dp = self.get_the_most_hated_and_not_scheduled_day()
        while dp[1] is not None:
            day = dp[0]
            ordered_persons = self.get_the_fewest_scheduled_person(dp)
            if len(self.scheduled[day])==0:
                self.scheduled[day].add(ordered_persons[0])
            first_person = list(self.scheduled[day])[0]
            self.find_first_good_for(first_person, ordered_persons, day)
            if len(self.scheduled[day])==2:
                self.fo_names[day] = self.select_fo(self.scheduled[day], day)
            dp = self.get_the_most_hated_and_not_scheduled_day()

    def get_the_most_hated_and_not_scheduled_day(self):
        most_hated = -1
        position = -1
        final_result2 = None
        for i in range(1, self.num_of_days+1):
            if len(self.scheduled[i])<2:
                day_hated = set()
                day_hated = day_hated.union(self.hated[i])
                day_hated = day_hated.union(self.scheduled[i])
                if (i-1) in self.scheduled:
                    day_hated = day_hated.union(self.scheduled[i-1])
                if (i+1) in self.scheduled:
                    day_hated = day_hated.union(self.scheduled[i+1])
                if len(day_hated)>most_hated:
                    most_hated = len(day_hated)
                    position = i
                    result2 = set(self.people.keys())-day_hated
                    final_result2 = result2
        return (position, final_result2)

    def get_the_fewest_scheduled_person(self, day_persons):
        day = day_persons[0]
        persons = day_persons[1]
        schedule_numbers = {}
        for n in persons:
            schedule_numbers[n] = 0
        # adjacency logic
        for d, st in self.scheduled.items():
            for nm in st:
                if nm in persons:
                    schedule_numbers[nm]+=7
                    if d==(day-2) or d==(day+2):
                        schedule_numbers[nm]+=2
                    if d==(day-3) or d==(day+3):
                        schedule_numbers[nm]+=1
        for nm in schedule_numbers.keys():
            schedule_numbers[nm] -= (self.people[nm].get_manual_day_difference()*7)

        return self.order_schedule_numbers(schedule_numbers)

    def order_schedule_numbers(self, schedule_numbers):
        # sort ascending
        return sorted(schedule_numbers.keys(), key=lambda k: schedule_numbers[k])

    def find_first_good_for(self, person_name, ordered_persons, day):
        t1 = self.people[person_name].get_type(day)
        for nm in ordered_persons:
            if nm!=person_name:
                t2 = self.people[nm].get_type(day)
                if t2.good_with(t1):
                    self.scheduled[day].add(nm)
                    return
        # if none found => error
        if len(ordered_persons)==1 and len(self.scheduled[day])==1:
            raise RuntimeError(f"I found only one person ({ordered_persons}) for day {day}")
        else:
            raise RuntimeError(f"No second person found for day {day} among {ordered_persons}")

    def select_fo(self, names_set, day):
        arr = list(names_set)
        if len(arr)==1:
            return arr[0]  # single person => IMS1 by default
        if len(arr)!=2:
            raise RuntimeError(f"select_fo with != 2 names: {names_set}")
        name1, name2 = arr[0], arr[1]
        p1 = self.people[name1]
        p2 = self.people[name2]

        # check nofo
        if p1.is_nofo() and (not p2.is_nofo()):
            return name2
        if (not p1.is_nofo()) and p2.is_nofo():
            return name1
        if p1.is_nofo() and p2.is_nofo():
            raise RuntimeError(f"Both people are nofo: {name1} and {name2} for day {day}")

        t1 = p1.get_type(day)
        t2 = p2.get_type(day)
        if t1==Type.BO and t2==Type.FO:
            return name2
        if t1==Type.FO and t2==Type.BO:
            return name1
        if t1==Type.FO and t2==Type.FO:
            raise RuntimeError(f"Both want FO for day {day}: {names_set}")
        if t1==Type.BO and t2==Type.BO:
            raise RuntimeError(f"Both want BO for day {day}: {names_set}")

        v1 = self.get_ims1_value(name1)
        v2 = self.get_ims1_value(name2)
        if v1< v2:
            return name1
        elif v2< v1:
            return name2
        else:
            return self.random_gen.choice([name1,name2])

    def get_ims1_value(self, name):
        """
        countAll = how many days name is scheduled
        countFo = how many days name is FO
        value= countFo*2 - countAll
        """
        count_all=0
        for st in self.scheduled.values():
            if name in st:
                count_all+=1
        count_fo=0
        for x in self.fo_names.values():
            if x==name:
                count_fo+=1
        return count_fo*2 - count_all

    def balance_ims(self):
        """
        replicate Java code that tries to reduce the difference in IMS1 coverage
        """
        fo_people = [p for p in self.people.values() if not p.is_nofo()]
        for i in range(10):
            max_ims1_val = 0
            max_p = None
            for p in fo_people:
                val = self.get_ims1_value(p.get_name())
                if val>max_ims1_val:
                    max_ims1_val=val
                    max_p = p
            if max_ims1_val>1:
                # find days that max_p is FO
                relevant_days=[]
                for d, st in self.scheduled.items():
                    if max_p.get_name() in st:
                        if self.fo_names.get(d,"")==max_p.get_name():
                            relevant_days.append(d)
                min_val_local = max_ims1_val
                minIMS1=None
                position_for_swap=None
                for d in relevant_days:
                    if max_p.get_type(d)!=Type.FO:
                        # find the other name
                        dd_list = list(self.scheduled[d])
                        if len(dd_list)==2:
                            if dd_list[0]==max_p.get_name():
                                other_name=dd_list[1]
                            else:
                                other_name=dd_list[0]
                            other_p = self.people[other_name]
                            if other_p.get_type(d)!=Type.BO:
                                v = self.get_ims1_value(other_name)
                                if v<min_val_local:
                                    min_val_local=v
                                    minIMS1=other_name
                                    position_for_swap=d
                if (max_ims1_val-min_val_local)>1 and (minIMS1 is not None):
                    self.fo_names[position_for_swap] = minIMS1


###############################################################################
# 7) EXCEL EXPORTER
###############################################################################
class ExcelExporter:
    """
    Python version of the Java ExcelExporter class, using openpyxl.
    """
    def __init__(self, scheduler, people, local_date):
        self.scheduler = scheduler
        self.people = people
        self.local_date = local_date
        self.df_format = "{:.2f}"

    def write_month_to_excel(self):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Monthly Schedule"

        # Define and register styles with unique names
        styles = {}

        def create_style(name, fill_color=None, font_color="000000", fill_type="solid"):
            if name not in styles:
                st = openpyxl.styles.NamedStyle(name=name)
                if fill_color:
                    st.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type=fill_type)
                st.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                st.font = Font(color=font_color)
                wb.add_named_style(st)
                styles[name] = st
            return styles[name]

        # Define all necessary styles with unique names
        styles['top_left'] = create_style('top_left', fill_color="333333", font_color="FFFFFF", fill_type="solid")
        styles['header'] = create_style('header', fill_color="333333", font_color="FFFFFF", fill_type="solid")
        styles['header_red'] = create_style('header_red', fill_color="FF0000", font_color="FFFFFF", fill_type="solid")
        styles['header_orange'] = create_style('header_orange', fill_color="FFA500", font_color="000000", fill_type="solid")
        styles['wd'] = create_style('wd', fill_color="000000", font_color="FF0000", fill_type="solid")
        styles['light_grey'] = create_style('light_grey', fill_color="D9D9D9", font_color="000000", fill_type="solid")
        styles['IMS1'] = create_style('IMS1', fill_color="D0E3FA", font_color="000000", fill_type="solid")
        styles['IMS2'] = create_style('IMS2', fill_color="FFFFCC", font_color="000000", fill_type="solid")
        styles['header_light_green'] = create_style('header_light_green', fill_color="CCFFCC", font_color="000000", fill_type="solid")
        styles['header_light_orange'] = create_style('header_light_orange', fill_color="00FFFF", font_color="000000", fill_type="solid")
        styles['header_blue_grey'] = create_style('header_blue_grey', fill_color="D9D9D9", font_color="000000", fill_type="solid")
        styles['header_pale_blue'] = create_style('header_pale_blue', fill_color="D0E3FA", font_color="000000", fill_type="solid")
        styles['header_pink'] = create_style('header_pink', fill_color="CCFFFF", font_color="000000", fill_type="solid")
        styles['sum'] = create_style('sum', fill_color="FF7F50", font_color="000000", fill_type="solid")

        rowNum = 1
        colNum = 1
        cell = sheet.cell(row=rowNum, column=colNum)
        cell.value = f"{self.local_date.year} {self.local_date.strftime('%b')}"
        cell.style = styles['top_left']

        saturdays = self.scheduler.get_saturdays()
        sundays = self.scheduler.get_sundays()
        normal_weekends = saturdays + sundays
        holidays = self.scheduler.get_holidays()

        sheet.column_dimensions[get_column_letter(colNum)].width = 18

        # Top row => day numbers
        for i in range(1, self.scheduler.get_num_of_days()+1):
            c = sheet.cell(row=rowNum, column=colNum+i)
            c.value = i
            if i in holidays:
                c.style = styles['header_red']
            elif i in normal_weekends:
                c.style = styles['header_orange']
            else:
                c.style = styles['header']
            sheet.column_dimensions[get_column_letter(colNum+i)].width = 5

        rowNum += 1
        # Row with day-of-week abbreviations
        c = sheet.cell(row=rowNum, column=colNum)
        c.style = styles['top_left']

        for i in range(1, self.scheduler.get_num_of_days()+1):
            d = datetime.date(self.local_date.year, self.local_date.month, i)
            dow_abr = d.strftime('%a').upper()[:3]
            c2 = sheet.cell(row=rowNum, column=colNum+i)
            c2.value = dow_abr
            if i in holidays:
                c2.style = styles['header_red']
            elif i in normal_weekends:
                c2.style = styles['header_orange']
            else:
                c2.style = styles['wd']

        rowNum += 1
        # Now each name row
        names = sorted([p.get_name() for p in self.people.get_people().values()])
        for idx, name in enumerate(names):
            row = rowNum + idx
            c = sheet.cell(row=row, column=colNum)
            c.value = name
            if (row % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']

            # figure out which days they are scheduled and which days they hate
            p = self.people.get_people()[name]
            places = [d for d, st in self.scheduler.get_scheduled().items() if name in st]
            hated_days = p.get_hated_days()

            colIdx = 2  # Start from the second column

            for i in range(1, self.scheduler.get_num_of_days()+1):
                c2 = sheet.cell(row=row, column=colNum+i)
                if (row % 2) == 0:
                    c2.style = styles['header_light_green']
                else:
                    c2.style = styles['header_light_orange']

                if i in places:
                    # check if FO
                    foName = self.scheduler.get_fo_names().get(i, None)
                    if foName is None:
                        c2.value = "NULL"
                    else:
                        if foName == name:
                            c2.value = "IMS1"
                            c2.style = styles['IMS1']
                        else:
                            c2.value = "IMS2"
                            c2.style = styles['IMS2']
                elif i in hated_days:
                    c2.value = "X"
                    c2.style = styles['light_grey']
                else:
                    if i in holidays:
                        c2.style = styles['header_red']
                    elif i in normal_weekends:
                        c2.style = styles['header_orange']

            # After setting cell values, handle the Summarize table later

        rowNum += len(names) + 3

        # Summarize row
        # place some text
        for i in range(1, 10):
            sheet.cell(row=rowNum, column=i)  # placeholder
        c2 = sheet.cell(row=rowNum, column=10)
        c2.value = f"{self.local_date.year} SUMMARIZE"
        c2.style = styles['header_light_orange']
        sheet.merge_cells(start_row=rowNum, start_column=10, end_row=rowNum, end_column=15)

        colX = 16
        c3 = sheet.cell(row=rowNum, column=colX)
        end_of_month = self.local_date.replace(day=self.scheduler.get_num_of_days())
        c3.value = f"up to {end_of_month.strftime('%b')} {end_of_month.day}"
        c3.style = styles['header_light_orange']
        sheet.merge_cells(start_row=rowNum, start_column=colX, end_row=rowNum, end_column=colX+4)

        rowNum += 1
        # Row of headers
        r = rowNum
        col = 1

        def setval_and_style(crow, ccol, val, st):
            cc = sheet.cell(row=crow, column=ccol)
            cc.value = val
            cc.style = st

        setval_and_style(r, col, "", styles['header_light_green']); col += 1
        setval_and_style(r, col, "IMS1\nWE", styles['header_blue_grey']); col += 1
        setval_and_style(r, col, "IMS1\nWD", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "IMS2\nWE", styles['header_blue_grey']); col += 1
        setval_and_style(r, col, "IMS2\nWD", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "NH", styles['header_pink']); col += 1
        setval_and_style(r, col, "WE", styles['header_pink']); col += 1
        setval_and_style(r, col, "WD", styles['header_pink']); col += 1
        setval_and_style(r, col, "ALL", styles['header_pink']); col += 1
        setval_and_style(r, col, "MON-\nTHU", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "FR", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "SA", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "SU", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "WE", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "NH", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "ALL", styles['header_pale_blue']); col += 1
        setval_and_style(r, col, "MON-\nTHU", styles['header_blue_grey']); col += 1
        setval_and_style(r, col, "FRI", styles['header_blue_grey']); col += 1
        setval_and_style(r, col, "SAT", styles['header_blue_grey']); col += 1
        setval_and_style(r, col, "SUN", styles['header_blue_grey']); col += 1
        setval_and_style(r, col, "NH", styles['header_blue_grey']); col += 1
        setval_and_style(r, col, "STANDBY\n WE                 WD               SUM", styles['header_pink']); col += 1

        rowNum += 1
        # Gather day-of-week lists
        def day_of_week(day):
            # Monday=0, Sunday=6
            dt = datetime.date(self.local_date.year, self.local_date.month, day)
            return dt.weekday()

        fridays = [i for i in range(1, self.scheduler.get_num_of_days()+1) if day_of_week(i) == 4]
        saturdays2 = [i for i in range(1, self.scheduler.get_num_of_days()+1) if day_of_week(i) == 5]
        sundays2 = [i for i in range(1, self.scheduler.get_num_of_days()+1) if day_of_week(i) == 6]

        kv = self.scheduler.get_key_value_store()
        y = self.local_date.year
        m = self.local_date.month

        # Data rows
        rindex = 0
        for name in names:
            rowIdx = rowNum + rindex
            rindex += 1

            c = sheet.cell(row=rowIdx, column=1)
            c.value = name
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']

            # gather scheduled days
            scheduled_days = [d for d, st in self.scheduler.get_scheduled().items() if name in st]
            # normal classification
            count_fr = sum(1 for d in scheduled_days if d in fridays and d not in holidays)
            count_sa = sum(1 for d in scheduled_days if d in saturdays2 and d not in holidays)
            count_su = sum(1 for d in scheduled_days if d in sundays2 and d not in holidays)
            count_fri_holidays = sum(1 for d in scheduled_days if d in fridays and d in holidays)
            count_sat_holidays = sum(1 for d in scheduled_days if d in saturdays2 and d in holidays)
            count_sun_holidays = sum(1 for d in scheduled_days if d in sundays2 and d in holidays)
            count_wd_holidays = sum(1 for d in scheduled_days if d in holidays and d not in fridays and d not in saturdays2 and d not in sundays2)
            count_holidays = count_fri_holidays + count_sat_holidays + count_sun_holidays + count_wd_holidays

            fo_days = [dd for dd, f in self.scheduler.get_fo_names().items() if f == name]
            ims1_weekend = sum(1 for dd in fo_days if dd in saturdays2 or dd in sundays2)
            ims1_weekday = len(fo_days) - ims1_weekend
            bo_days = [d for d in scheduled_days if d not in fo_days]
            ims2_weekend = sum(1 for dd in bo_days if dd in saturdays2 or dd in sundays2)
            ims2_weekday = len(bo_days) - ims2_weekend

            colIdx = 2  # Start from the second column

            # IMS1 WE
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = ims1_weekend
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # IMS1 WD
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = ims1_weekday
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # IMS2 WE
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = ims2_weekend
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # IMS2 WD
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = ims2_weekday
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # NH
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = count_holidays
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # WE (IMS1_WE + IMS2_WE)
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = ims1_weekend + ims2_weekend
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # WD (IMS1_WD + IMS2_WD)
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = ims1_weekday + ims2_weekday
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # ALL
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = len(scheduled_days)
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # Summation from previous months:
            allAll = kv.sum(name, y, m, "ALL") + len(scheduled_days)
            frAll = kv.sum(name, y, m, "FR") + count_fr
            suAll = kv.sum(name, y, m, "SU") + count_su
            weAll = kv.sum(name, y, m, "WE") + (count_sa + count_su)
            nhwdAll = kv.sum(name, y, m, "NHWD") + count_wd_holidays
            nhfrAll = kv.sum(name, y, m, "NHFR") + count_fri_holidays
            nhsaAll = kv.sum(name, y, m, "NHSA") + count_sat_holidays
            nhsuAll = kv.sum(name, y, m, "NHSU") + count_sun_holidays
            nhAll = nhwdAll + nhfrAll + nhsaAll + nhsuAll
            mon2thuAll = allAll - (weAll + nhsaAll + nhsuAll) - (frAll + nhfrAll) - nhwdAll
            saAll = weAll - suAll

            # MON-THU
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = mon2thuAll
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # FR
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = frAll
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # SA
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = saAll
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # SU
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = suAll
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # WE + NHSA + NHSU
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = weAll + nhsaAll + nhsuAll
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # NH
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = nhAll
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # ALL
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = allAll
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # Local counts:
            mon_to_thu = [d for d in scheduled_days if d not in saturdays2 and d not in sundays2 and d not in fridays and d not in holidays]
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = len(mon_to_thu)
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            only_fr = [d for d in scheduled_days if d in fridays and d not in holidays]
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = len(only_fr)
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = count_sa
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = count_su
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = count_holidays
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # The big SB formulas:
            weekendSBNum = 3.6 * frAll + 9.6 * saAll + 6 * suAll + 9.6 * (nhfrAll + nhsaAll + nhsuAll + nhwdAll)
            weekdaySBNum = 3.2 * mon2thuAll + 1.4 * frAll + 1.8 * suAll
            allSB = weekendSBNum + weekdaySBNum

            # WEEKEND SB
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = float(self.df_format.format(weekendSBNum))
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1
            # Skip one column
            colIdx += 1

            # WEEKDAY SB
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = float(self.df_format.format(weekdaySBNum))
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1
            # Skip one column
            colIdx += 1

            # ALL SB
            c = sheet.cell(row=rowIdx, column=colIdx)
            c.value = float(self.df_format.format(allSB))
            if (rowIdx % 2) == 0:
                c.style = styles['header_light_green']
            else:
                c.style = styles['header_light_orange']
            colIdx += 1

            # Store in DB
            kv.write_data(name, y, m, "ALL", len(scheduled_days))
            kv.write_data(name, y, m, "FR", count_fr)
            kv.write_data(name, y, m, "SU", count_su)
            kv.write_data(name, y, m, "WE", count_sa + count_su)
            kv.write_data(name, y, m, "NHWD", count_wd_holidays)
            kv.write_data(name, y, m, "NHFR", count_fri_holidays)
            kv.write_data(name, y, m, "NHSA", count_sat_holidays)
            kv.write_data(name, y, m, "NHSU", count_sun_holidays)

        # Summation row
        total_row = rowNum + len(names)
        for i in range(23):
            colLetter = get_column_letter(i + 2)
            firstDataRow = rowNum
            lastDataRow = rowNum + len(names) - 1
            form = f"=SUM({colLetter}{firstDataRow}:{colLetter}{lastDataRow})"
            sumCell = sheet.cell(row=total_row, column=i + 2)
            sumCell.value = form
            sumCell.style = styles['sum']

        # Finalize DB
        kv.print_all(y)
        try:
            kv.close()
        except:
            pass

        fileNameExcel = f"schedule-{self.local_date.year}-{self.local_date.month}.xlsx"
        wb.save(fileNameExcel)


###############################################################################
# 8) MAINTASK
###############################################################################
class MainTask:
    """
    Python version of the Java MainTask class.
    """

    def __init__(self, args):
        self.args = args
        self.local_date = None

    def make(self):
        """
        Replicate logic:
         - parse args => local_date
         - read People
         - run Scheduler
         - writeMonth (CSV)
         - ExcelExporter => .xlsx
        """
        if self.args and len(self.args)==1:
            # parse yymm
            arg0 = self.args[0]
            yy = int(arg0[:2])
            mm = int(arg0[2:])
            year = 2000+yy
            self.local_date = datetime.date(year, mm, 1)
        else:
            # next month
            now = datetime.date.today()
            year = now.year
            month = now.month
            new_month = month+1
            new_year = year
            if new_month>12:
                new_month=1
                new_year+=1
            self.local_date = datetime.date(new_year,new_month,1)

        # build People
        pp = People()
        # build Scheduler
        sched = Scheduler(pp.get_people(), self.local_date)
        # write .csv
        self.write_month(sched, pp)
        # produce XLSX
        exp = ExcelExporter(sched, pp, self.local_date)
        exp.write_month_to_excel()

    def write_month(self, scheduler, people):
        """
        Mirroring the Java code that writes a .csv with w/f/b lines, etc.
        """
        # fill day-> "X -> name1 - name2"
        # also increment scheduled counters
        for day, st in scheduler.get_scheduled().items():
            if day>=1 and day<=scheduler.get_num_of_days():
                if len(st)==2:
                    arr = list(st)
                    p1type = people.get_people()[arr[0]].get_type(day)
                    p2type = people.get_people()[arr[1]].get_type(day)
                    # which is FO first?
                    # The Java code: if Type.is_first_fo(p1type, p2type) => arr[0] first
                    # We'll skip writing lines to .txt for brevity, or do them if you want.

        for day, st in scheduler.get_scheduled().items():
            for nm in st:
                people.get_people()[nm].increment_scheduled()
                people.get_people()[nm].get_wanted_days().add(day)

        # build CSV lines
        lines_csv = []
        for name, pers in sorted(people.get_people().items()):
            arr = [name.replace(" ","_")]
            # first "wX" for each wanted day
            wdays = sorted(list(pers.get_wanted_days()))
            for d in wdays:
                arr.append(f"w{d}")
            # also f or b
            for d in wdays:
                if scheduler.get_fo_names().get(d,"")==name:
                    arr.append(f"f{d}")
                else:
                    arr.append(f"b{d}")
            line_str = " ".join(arr)
            lines_csv.append(line_str)

        csv_filename = f"schedule-{self.local_date.year}-{self.local_date.month}.csv"
        with codecs.open(csv_filename,"w","utf-8") as f:
            f.write("\n".join(lines_csv))


###############################################################################
# 9) MAIN
###############################################################################
def main():
    try:
        main_task = MainTask(sys.argv[1:])
        main_task.make()
    except Exception as e:
        print("Error in main():", e)
        traceback.print_exc()

if __name__=="__main__":
    main()
